# coding: utf-8
import os
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'invoice'
ws['A1'] = 'T-shirt'
ws.cell(row=1, column=2, value=10)

path_book = 'excel'
if not os.path.isdir(path_book):
    os.makedirs(path_book)

wb.save(path_book + '/text.xlsx')
