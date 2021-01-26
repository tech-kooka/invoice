# coding: utf-8
import os
from openpyxl import Workbook

wb = Workbook()
# grab the active worksheet
ws = wb.active


# Data can be assigned directly to cells
ws['B2'] = 'ABC Co.'

ws['B3'] = '東京都新宿区西早稲田2-3-18'
ws['B4'] = 'abc.co.jp'
ws['B5'] = '03-2222-4444'

ws['B9'] = 'BILL TO'
ws['B10'] = '<Contact Name>'
ws['B11'] = '<Client Company Name>'
ws['B12'] = '<Address>'
ws['B13'] = '<Phone, Email>'

ws['D9'] = 'SHIP TO'
ws['D10'] = '<Name / Dept>'
ws['D11'] = '<Client Company Name>'
ws['D12'] = '<Address>'
ws['D13'] = '<Phone, Email>'

ws['E2'] = 'INVOICE'

invoice_no = '#INV00001'
invoice_day = '2020/12/01'
due_day = '2021/01/31'

ws['E9'] = 'Invoice No:'
ws['F9'] = invoice_no
ws['E10'] = 'Invoice Date:'
ws['F10'] = invoice_day
ws['E11'] = 'Due Date:'
ws['F11'] = due_day

ws['B15'] = 'DESCRIPTION'
ws['D15'] = 'QTY'
ws['E15'] = 'UNIT PRICE'
ws['F15'] = 'TOTAL'

ws['E27'] = 'SUBTOTAL'
ws['F27'] = '=SUM(F16:F26)'

ws['E28'] = 'TAX RATE'
ws['F28'] = 0.10

ws['E29'] = 'TOTAL TAX'
ws['F29'] = '=F27*F28'

ws['E30'] = 'SHIPPING/HANDLING'
ws['F30'] = 0

ws['E31'] = 'Balance Due'
ws['F31'] = '=F27+F29+F30'

ws['B28'] = 'Tank you for your business!'


# Save the file

path_book = 'excel'
if not os.path.isdir(path_book):
    os.makedirs(path_book)

wb.save(path_book + "/invoice_" + invoice_no + ".xlsx")