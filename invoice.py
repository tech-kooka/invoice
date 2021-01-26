# coding: utf-8
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image


def createInvoice(ws, invoiceDate, dueDate, invoiceNo):

    ws.sheet_view.showGridLines = False

    # Data can be assigned directly to cells
    ws['B2'] = 'ABC Co.'

    ws['B3'] = '東京都新宿区西早稲田2-3-18'
    ws['B4'] = 'abc.co.jp'
    ws['B5'] = '03-2222-4444'

    ws['E2'] = 'INVOICE'

    invoice_no = '#INV00001'

    ws['E9'] = 'Invoice No:'
    ws['F9'] = invoiceNo
    ws['E10'] = 'Invoice Date:'
    ws['F10'] = invoiceDate
    ws['E11'] = 'Due Date:'
    ws['F11'] = dueDate

    ws['F10'].number_format = 'yyyy年m月d日'
    ws['F11'].number_format = 'yyyy年m月d日'

    ws['B15'] = 'DESCRIPTION'
    ws['D15'] = 'QTY'
    ws['E15'] = 'UNIT PRICE'
    ws['F15'] = 'TOTAL'

    ws['E27'] = 'SUBTOTAL'
    ws['F27'] = '=SUM(F16:F26)'

    ws['E28'] = 'TAX RATE'
    ws['F28'] = 0.10

    ws['E29'] = 'TOTAL TAX'
    ws['F29'] = '=F27*F28'

    ws['E30'] = 'SHIPPING/HANDLING'
    ws['F30'] = 0

    ws['E31'] = 'Balance Due'
    ws['F31'] = '=F27+F29+F30'

    ws['B28'] = 'Tank you for your business!'

    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    ws['F27'].number_format = '#,##0'
    ws['F28'].number_format = '0%'
    ws['F30'].number_format = '#,##0'
    ws['F31'].number_format = '#,##0'

    # Font
    font = Font(name='MS PGothic')
    # # write in sheet
    for row in ws:
        for cell in row:
            ws[cell.coordinate].font = font

    ws['B2'].font = Font(size=20, bold=True)
    ws['E31'].font = Font(size=18, bold=True)
    ws['F31'].font = Font(size=18, bold=True)
    ws['E2'].font = Font(size=21, bold=True)

    #alignment
    rng = ws['E9':'E11']
    for row in rng:
        for c in row:
            c.alignment = Alignment(horizontal = 'right', vertical = 'center',wrap_text = False)
    rng = ws['B15':'F15']
    for col in rng:
        for r in col:
            r.alignment = Alignment(horizontal = 'center', vertical = 'center',wrap_text = False)
    rng = ws['E27':'E33']
    for row in rng:
        for c in row:
            c.alignment = Alignment(horizontal = 'right', vertical = 'center',wrap_text = False)
    
    ws['F9'].alignment = Alignment(horizontal = 'right', vertical = 'center',wrap_text = False)
    ws['E2'].alignment = Alignment(horizontal = 'right', vertical = 'center',wrap_text = False)
    ws['B28'].alignment = Alignment(horizontal = 'center', vertical = 'center',wrap_text = False)

    #Border
    ws['B9'].border = Border(bottom=Side(style='thin', color='E9EEEB'))
    ws['D9'].border = Border(bottom=Side(style='thin', color='E9EEEB'))
    ws['E31'].border = Border(bottom=Side(style='thick', color='666666'))
    ws['F31'].border = Border(bottom=Side(style='thick', color='666666'))

    rng = ws['B15':'F15']
    for col in rng:
        for r in col:
            r.alignment = Alignment(horizontal = 'center', vertical = 'center',wrap_text = False)
            r.border = Border(top=Side(style='thin', color='E9EEEB'), 
                    bottom=Side(style='thin', color='E9EEEB'), 
                    left=Side(style='thin', color='E9EEEB'),
                    right=Side(style='thin', color='E9EEEB')
                )

    rng = ws['B16':'F26']
    for row in range(16, 27):
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = Border(top=Side(style='thin', color='E9EEEB'), 
                    bottom=Side(style='thin', color='E9EEEB'), 
                    left=Side(style='thin', color='E9EEEB'),
                    right=Side(style='thin', color='E9EEEB')
                )
            if row % 2 != 0:
                ws.cell(row=row, column=col).fill = PatternFill(patternType='solid',
                                    fgColor='E0FFE0', bgColor='E0FFE0')
            if col == 6:
                ws.cell(row=row, column=col).value ='=D'+ str(row) + '*E' + str(row)
                ws.cell(row=row, column=col).number_format = '#,##0'

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
    ws.merge_cells(start_row=28, start_column=2, end_row=32, end_column=3)

    img = Image('abc.png')
    img.width = 100
    img.height = 100
    ws.add_image(img, 'F3')