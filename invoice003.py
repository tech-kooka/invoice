# coding: utf-8
import os
from datetime import datetime, timedelta, date
from openpyxl import Workbook
import pandas as pd

from invoice import createInvoice


#date
invoice_day = datetime.now()
year = invoice_day.year
month = invoice_day.month
if invoice_day.month == 12:
    year = invoice_day.year + 1
    month = 0
due_day = date(year, month +  1 ,1) - timedelta(days=1)

##invoice data
df_invoice = pd.read_csv('invoice_data.csv')
df_invoice['date'] = pd.to_datetime(df_invoice['date'])
df_invoice = df_invoice[(df_invoice['date'] >= pd.to_datetime(date(invoice_day.year, invoice_day.month ,1))) & (df_invoice['date'] <= pd.to_datetime(due_day))]
#print(df_invoice)

company_invioce = sorted(list(df_invoice['company'].unique()))

## company data
df_company = pd.read_csv('company.csv')


# create invoice
for company in company_invioce:

    wb = Workbook()
    # grab the active worksheet
    ws = wb.active

    invoice_no = company + '-' + str(invoice_day.month)
    createInvoice(ws, invoice_day, due_day, invoice_no)

    # get company info
    df_cominfo = df_company[df_company['id'] == company]
    # print('company' + df_cominfo) 
    

    ws['B9'] = 'BILL TO'
    ws['B10'] = df_cominfo.iloc[0, 2]
    ws['B11'] = df_cominfo.iloc[0, 3]
    ws['B12'] = df_cominfo.iloc[0, 4]
    ws['B13'] = df_cominfo.iloc[0, 5] + ' ' + df_cominfo.iloc[0, 6]

    ws['D9'] = 'SHIP TO'
    ws['D10'] = df_cominfo.iloc[0, 7]
    ws['D11'] = df_cominfo.iloc[0, 8]
    ws['D12'] = df_cominfo.iloc[0, 9]
    ws['D13'] = df_cominfo.iloc[0, 10] + ' ' + df_cominfo.iloc[0, 11]

    # get invoice data
    df_invoiceData = df_invoice[df_invoice['company'] == company ]
    values = df_invoiceData.values.tolist()
    # print(values)

    # rng = ws['B16':'F26']
    y = 0
    for record in values:
        y += 1
        for x, data in enumerate(record):
            col = 2+x
            row = 15+y
            if x == 1:
                ws.cell(row=row, column=2, value=data)
            elif x == 2:
                ws.cell(row=row, column=4, value=data)
            elif x == 3:
                ws.cell(row=row, column=5, value=data)


        # for col in range(2, 7):
        #     ws.cell(row=row, column=col, value=1)
    
    # Save the file
    path_book = 'excel'
    if not os.path.isdir(path_book):
        os.makedirs(path_book)

    wb.save(path_book + "/" + company + "invoice_" + invoice_no + ".xlsx")

    print('complate ' + company)





